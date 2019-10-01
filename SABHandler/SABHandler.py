import os
import xlrd
import openpyxl as excel
from openpyxl import Workbook, load_workbook


# Ham chuyen doi file xls sang xlsx
def cvt_xls_to_xlsx(src_file_path, name_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0, len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(name_file_path + ".xlsx")


# Initial Input data from SAB file
#input: xlsx file
def initDataInput(url, name_file):
    out_data_path = "_int_" + name_file + ".xlsx"
    #
    print("msg Checking file SAB path...")
    if os.path.exists(url):
        print("msg SAB found!")
        print("msg Initial Input data from SAB file")

        # open GCAPE_Gen1.x_SAB_Interface_Gen_B file
        sab_wb = excel.load_workbook(filename=url)
        sab_ws = sab_wb['Interfaces']

        # create initDataInput file
        # SAB interface Sheet
        data_output_wb = Workbook()
        data_output_ws = data_output_wb.active
        data_output_ws.title = "Interfaces"

        for row in sab_ws:
            for cell in row:
                data_output_ws[cell.coordinate].value = cell.value

        data_output_ws.delete_rows(1, 8)
        data_output_ws.delete_cols(1)
        data_output_wb.save(out_data_path)

        # create AutoTypeDefinition sheet
        sab_ws = sab_wb['AutoTypeDefinition']
        data_output_ws = data_output_wb.create_sheet("AutoTypeDefinition")

        for row in sab_ws:
            for cell in row:
                data_output_ws[cell.coordinate].value = cell.value

        data_output_ws.delete_rows(1, 10)
        data_output_wb.save(out_data_path)

        sab_wb.close()
        data_output_wb.close()

        print("initDataInput is Done!")

    else:
        print("error File SAB not found")

title_door = ("Component","Object Text","SAB Interface","H_SafetyClassification","TS_Module_Name","TS_Test Case Status","TS_Test Type","TS_Test Priority","TS_Object type","TS_Precondition / Dependency","TS_Test Description","TS_Expected Result", "Return", "Interface Type")

f_Component = 1
f_Object_Text = 2
f_SAB_Interface = 3
f_H_SafetyClassification = 4
f_TS_Module_Name = 5
f_TS_Test_Case_Status = 6
f_TS_Test_Type = 7
f_TS_Test_Priority = 8
f_TS_Object_type = 9
f_TS_Precondition_Dependency = 10
f_TS_Test_Description = 11
f_TS_Expected_Result = 12
f_Return = 13

def readDataInput(data_path, name_file):
    wb = load_workbook(data_path)
    data = wb['Interfaces']
    max_row = data.max_row
    max_column = data.max_column

    # output file
    output = Workbook()
    tc = output.active
    tc.title = "PreOutput"

    # output file
    wb2 = Workbook()
    ListInterface = wb2.create_sheet("ListInterfaceSR_NotReturn", 0)

    for i in range(1, max_column + 1):
        ListInterface.cell(row=1, column=i).value = data.cell(row=1, column=i).value
    for i in range(0, 14):
        tc.cell(row=1, column=i + 1).value = title_door[i]

    # --------------
    # Auto generation test case
    rows = 2
    lists_row = 2
    for data_rows in range(2, max_row + 1):
        # ----------------
        # Because for Event and Timer is not support now:
        # It will be define later
        kiemtra = data.cell(row=data_rows, column=4).value
        if kiemtra == "Event" or kiemtra == "Timer":
            continue
        else:
            # 13. Return
            tc.cell(row=rows, column=f_Return).value = data.cell(row=data_rows, column=6).value
            # 14.interface type
            tc.cell(row=rows, column=14).value = data.cell(row=data_rows, column=4).value

            # 1.Component
            tc.cell(row=rows, column=f_Component).value = data.cell(row=data_rows, column=1).value

            # ----------------
            # 3.SAB Interface
            tc.cell(row=rows, column=f_SAB_Interface).value = data.cell(row=data_rows, column=30).value

            # ----------------
            # 4.H_SafetyClassification
            tc.cell(row=rows, column=f_H_SafetyClassification).value = data.cell(row=data_rows, column=29).value

            # ----------------
            # 5.TS_Module_Name
            tc.cell(row=rows, column=f_TS_Module_Name).value = "Interfaces"

            # ----------------
            # 6.TS_Test Case Status
            tc.cell(row=rows, column=f_TS_Test_Case_Status).value = "ready for review"

            # ----------------
            # 7.TS_Test Type
            tc.cell(row=rows, column=f_TS_Test_Type).value = "Interface test"

            # ----------------
            # 9.TS_Object type
            tc.cell(row=rows, column=f_TS_Object_type).value = "test case"

            data_Object_Text = ""
            data_TS_Precondition_Dependency = ""
            data_TS_Test_Description = ""
            data_TS_Expected_Result = ""
            value_test_case = ""
            return_value_check = ""
            # ----------------
            if data.cell(row=data_rows, column=4).value == "SR":

                return_value_check = data.cell(row=data_rows, column=6).value
                if return_value_check is None:
                    # List interface have return is empty
                    for i in range(1, max_column + 1):
                        ListInterface.cell(row=lists_row, column=i).value = data.cell(row=data_rows, column=i).value

                    lists_row = lists_row + 1

                # Write test case
                data_TS_Precondition_Dependency = "1. SW run normal \r\n 2. Input all available values"
                if data.cell(row=data_rows, column=5).value == "in":
                    # KT 26/03/2019
                    # Change Set to Get
                    data_Object_Text = "SAB_Get_" + data.cell(row=data_rows, column=2).value + "_" + data.cell(
                        row=data_rows, column=3).value
                    data_TS_Test_Description = "1. Run to where " + data.cell(row=data_rows,
                                                                              column=1).value + "_Main call required SAB" + "\r\n" + "2. Check " + data.cell(
                        row=data_rows, column=11).value
                    data_TS_Expected_Result = "1. SW halt at where " + data.cell(row=data_rows,
                                                                                 column=1).value + "_Main call required SAB" + "\r\n" + "2. " + data.cell(
                        row=data_rows, column=11).value + " equal to required values"
                elif data.cell(row=data_rows, column=5).value == "out":
                    # KT 26/03/2019
                    # Change Get to Set
                    data_Object_Text = "SAB_Set_" + data.cell(row=data_rows, column=2).value + "_" + data.cell(
                        row=data_rows, column=3).value
                    data_TS_Test_Description = "1. Run to where " + data.cell(row=data_rows,
                                                                              column=1).value + "_Main call required SAB" + "\r\n" + "2. Check " + data.cell(
                        row=data_rows, column=11).value
                    data_TS_Expected_Result = "1. SW halt at where " + data.cell(row=data_rows,
                                                                                 column=1).value + "_Main call required SAB" + "\r\n" + "2. " + data.cell(
                        row=data_rows, column=11).value + " equal to required values"

            elif data.cell(row=data_rows, column=4).value == "CS":
                data_Object_Text = "SAB_Call_" + data.cell(row=data_rows, column=2).value + "_" + data.cell(row=data_rows,
                                                                                                        column=3).value
                data_TS_Precondition_Dependency = "1. SW run normal" + "\r\n" + "2. Input all available values"
                data_TS_Test_Description = "1. Set BP1 at " + data.cell(row=data_rows, column=1).value + " in " + data.cell(row=data_rows, column=1).value + "\r\n" \
                                           + "2. Set BP2 at " + data.cell(row=data_rows, column=3).value + " in " + data.cell(row=data_rows, column=2).value + "\r\n"\
                                           + "3. When software halt at BP1 set value param1 & param2 " + "\r\n"\
                                           + "4. Press 'Run until' check flow of software  " + "\r\n" \
                                           + "5. After exit Mapping, function check data return  " + "\r\n"

                data_TS_Expected_Result = "1. BP1 reached" + "\r\n" + "2. BP2 reached"

            elif data.cell(row=data_rows, column=4).value == "Message":
                data_TS_Precondition_Dependency = "1. SW run normal \r\n 2. Input all available values"
                if data.cell(row=data_rows, column=5).value == "send":
                    data_Object_Text = "Set_" + data.cell(row=data_rows, column=2).value + "_" + data.cell(
                        row=data_rows, column=3).value
                    data_TS_Test_Description = "1. Run to where " + data.cell(row=data_rows,
                                                                              column=1).value + "_Main call required SAB" + "\r\n" + "2. Check " + data.cell(
                        row=data_rows, column=11).value
                    data_TS_Expected_Result = "1. SW halt at where " + data.cell(row=data_rows,
                                                                                 column=1).value + "_Main call required SAB" + "\r\n" + "2. " + data.cell(
                        row=data_rows, column=11).value + " equal to required values"
                elif data.cell(row=data_rows, column=5).value == "receive":
                    data_Object_Text = "Get_" + data.cell(row=data_rows, column=2).value + "_" + data.cell(
                        row=data_rows, column=3).value
                    data_TS_Test_Description = "1. Run to where " + data.cell(row=data_rows,
                                                                              column=1).value + "_Main call required SAB" + "\r\n" + "2. Check " + data.cell(
                        row=data_rows, column=11).value
                    data_TS_Expected_Result = "1. SW halt at where " + data.cell(row=data_rows,
                                                                                 column=1).value + "_Main call required SAB" + "\r\n" + "2. " + data.cell(
                        row=data_rows, column=11).value + " equal to required values"
            '''
            elif data.cell(row=data_rows, column=4).value == "Event"
                data_TS_Precondition_Dependency = "1. SW run normal \r\n 2. Input all available values"
                if data.cell(row=data_rows, column=5).value == ""
            '''

            # ----------------
            # 2.Object Text
            tc.cell(row=rows, column=f_Object_Text).value = data_Object_Text

            # ----------------
            # 8.TS_Test Priority
            if data.cell(row=data_rows, column=29).value == 'QM':
                tc.cell(row=rows, column=f_TS_Test_Priority).value = "medium"

            elif data.cell(row=data_rows, column=29).value == "ASIL A":
                tc.cell(row=rows, column=f_TS_Test_Priority).value = "high"

            elif data.cell(row=data_rows, column=29).value == "ASIL B":
                tc.cell(row=rows, column=f_TS_Test_Priority).value = "high"

            # ----------------
            # 10.TS_Precondition / Dependency
            tc.cell(row=rows, column=f_TS_Precondition_Dependency).value = data_TS_Precondition_Dependency

            # ----------------
            # 11.TS_Test Description
            tc.cell(row=rows, column=f_TS_Test_Description).value = data_TS_Test_Description

            # ----------------
            # 12.TS_Expected Result
            tc.cell(row=rows, column=f_TS_Expected_Result).value = data_TS_Expected_Result

            rows = rows + 1

    list_interface = "_List_Report_" + name_file + ".xlsx"
    wb2.save(list_interface)
    pre_output_path = "_preOut_" + name_file + ".xlsx"
    output.save(pre_output_path)

    data = wb['AutoTypeDefinition']
    tc = output.create_sheet("AutoTypeDefinition")
    for row in data:
        for cell in row:
            tc[cell.coordinate].value = cell.value

    output.save(pre_output_path)
    wb.close()
    wb2.close()
    output.close()
    print("Finished Create Test Case")



title_door_final = ("Component","Object Text","SAB Interface","H_SafetyClassification","TS_Module_Name","TS_Test Case Status","TS_Test Type","TS_Test Priority","TS_Object type","TS_Precondition / Dependency","TS_Test Description","TS_Expected Result")

# Xu ly du lieu voi Value_t refer to AutoTypeDefinition sheet
def FinalOutput(pre_output_path, name_file):
    # Open PreOutput.xlsx
    data = load_workbook(pre_output_path)
    in_data = data['PreOutput']
    in_type = data['AutoTypeDefinition']
    max_row_data = in_data.max_row
    max_row_type = in_type.max_row

    # create test case output
    out = Workbook()
    output = out.create_sheet("TestCaseOutput", 0)

    # List all AutoType is not defined in AutoTypeDefinition sheet
    list1 = out.create_sheet("ListTypeNotDef", 1)
    list1.cell(row=1, column=1).value = "Name of Type"

    # List all AutoType is not range
    list2 = out.create_sheet("ListTypeNotRange", 2)
    list2.cell(row=1, column=1).value = "Name of Type"

    # Coppy name of title for TestCaseOutput
    for i in range(0, 12):
        output.cell(row=1, column=i + 1).value = title_door_final[i]

    out_row = 2
    out_row2 = 2
    out_row3 = 2

    for data_row in range(2, max_row_data + 1):
        if in_data.cell(row=data_row, column=14).value == "SR":
            if in_data.cell(row=data_row, column=13).value is None:
                # Test case with SR but the value of return is None/not defined -> Will be create one test case with value (0)
                output.cell(row=out_row, column=1).value = in_data.cell(row=data_row, column=1).value
                output.cell(row=out_row, column=2).value = in_data.cell(row=data_row, column=2).value + " (0)"
                output.cell(row=out_row, column=3).value = in_data.cell(row=data_row, column=3).value
                output.cell(row=out_row, column=4).value = in_data.cell(row=data_row, column=4).value
                output.cell(row=out_row, column=5).value = in_data.cell(row=data_row, column=5).value
                output.cell(row=out_row, column=6).value = in_data.cell(row=data_row, column=6).value
                output.cell(row=out_row, column=7).value = in_data.cell(row=data_row, column=7).value
                output.cell(row=out_row, column=8).value = in_data.cell(row=data_row, column=8).value
                output.cell(row=out_row, column=9).value = in_data.cell(row=data_row, column=9).value
                output.cell(row=out_row, column=10).value = "1. SW run normal \r\n 2. Input = 0"
                output.cell(row=out_row, column=11).value = in_data.cell(row=data_row, column=11).value
                output.cell(row=out_row, column=12).value = in_data.cell(row=data_row, column=12).value + " (0)"
                out_row = out_row + 1
            else:
                type_defined = False
                range_defined = False
                list_value_range = []
                count_value_range = 0
                for type_row in range(2, max_row_type + 1):
                    # Type was defined in AutoTypeDefinition sheet
                    if in_type.cell(row=type_row, column=1).value == in_data.cell(row=data_row, column=13).value:
                        type_defined = True
                        if in_type.cell(row=type_row, column=4).value is not None:
                            # Range have value
                            range_defined = True
                            count_value_range = count_value_range + 1
                            list_value_range = list_value_range + [in_type.cell(row=type_row, column=4).value]

                # Write test case
                if (type_defined == False) or (range_defined == False):
                    # Write test case to TestCaseOutput
                    # Test case with SR but the return type is not defined or not value of range.-> Will be create one test case with value (0)
                    output.cell(row=out_row, column=1).value = in_data.cell(row=data_row, column=1).value
                    output.cell(row=out_row, column=2).value = in_data.cell(row=data_row, column=2).value + " (0)"
                    output.cell(row=out_row, column=3).value = in_data.cell(row=data_row, column=3).value
                    output.cell(row=out_row, column=4).value = in_data.cell(row=data_row, column=4).value
                    output.cell(row=out_row, column=5).value = in_data.cell(row=data_row, column=5).value
                    output.cell(row=out_row, column=6).value = in_data.cell(row=data_row, column=6).value
                    output.cell(row=out_row, column=7).value = in_data.cell(row=data_row, column=7).value
                    output.cell(row=out_row, column=8).value = in_data.cell(row=data_row, column=8).value
                    output.cell(row=out_row, column=9).value = in_data.cell(row=data_row, column=9).value
                    output.cell(row=out_row, column=10).value = "1. SW run normal \r\n 2. Input = 0"
                    output.cell(row=out_row, column=11).value = in_data.cell(row=data_row, column=11).value
                    output.cell(row=out_row, column=12).value = in_data.cell(row=data_row, column=12).value + " (0)"
                    out_row = out_row + 1

                    # Write Type is not define to ListTypeNotDef
                    if (type_defined == False):
                        list1.cell(row=out_row2, column=1).value = in_data.cell(row=data_row, column=13).value
                        out_row2 = out_row2 + 1

                # Ghi test case SR voi kieu tra ve duoc dinh nghia va range duoc dinh nghia
                # Tach cac gia tri cua range de tao test case.
                if (type_defined == True) and (range_defined == True):
                    if count_value_range == 1:
                        if (str(list_value_range[0]) == "0..255") or (str(list_value_range[0]) == "0...255"):
                            output.cell(row=out_row, column=1).value = in_data.cell(row=data_row, column=1).value
                            output.cell(row=out_row, column=2).value = in_data.cell(row=data_row,
                                                                                    column=2).value + " (0)"
                            output.cell(row=out_row, column=3).value = in_data.cell(row=data_row, column=3).value
                            output.cell(row=out_row, column=4).value = in_data.cell(row=data_row, column=4).value
                            output.cell(row=out_row, column=5).value = in_data.cell(row=data_row, column=5).value
                            output.cell(row=out_row, column=6).value = in_data.cell(row=data_row, column=6).value
                            output.cell(row=out_row, column=7).value = in_data.cell(row=data_row, column=7).value
                            output.cell(row=out_row, column=8).value = in_data.cell(row=data_row, column=8).value
                            output.cell(row=out_row, column=9).value = in_data.cell(row=data_row, column=9).value
                            output.cell(row=out_row, column=10).value = "1. SW run normal \r\n 2. Input = 0"
                            output.cell(row=out_row, column=11).value = in_data.cell(row=data_row, column=11).value
                            output.cell(row=out_row, column=12).value = in_data.cell(row=data_row,
                                                                                     column=12).value + " (0)"
                            out_row = out_row + 1
                            output.cell(row=out_row, column=1).value = in_data.cell(row=data_row, column=1).value
                            output.cell(row=out_row, column=2).value = in_data.cell(row=data_row,
                                                                                    column=2).value + " (255)"
                            output.cell(row=out_row, column=3).value = in_data.cell(row=data_row, column=3).value
                            output.cell(row=out_row, column=4).value = in_data.cell(row=data_row, column=4).value
                            output.cell(row=out_row, column=5).value = in_data.cell(row=data_row, column=5).value
                            output.cell(row=out_row, column=6).value = in_data.cell(row=data_row, column=6).value
                            output.cell(row=out_row, column=7).value = in_data.cell(row=data_row, column=7).value
                            output.cell(row=out_row, column=8).value = in_data.cell(row=data_row, column=8).value
                            output.cell(row=out_row, column=9).value = in_data.cell(row=data_row, column=9).value
                            output.cell(row=out_row, column=10).value = "1. SW run normal \r\n 2. Input = 255"
                            output.cell(row=out_row, column=11).value = in_data.cell(row=data_row, column=11).value
                            output.cell(row=out_row, column=12).value = in_data.cell(row=data_row,
                                                                                     column=12).value + " (255)"
                            out_row = out_row + 1
                        else:
                            output.cell(row=out_row, column=1).value = in_data.cell(row=data_row, column=1).value
                            output.cell(row=out_row, column=2).value = in_data.cell(row=data_row,
                                                                                    column=2).value + " (" + str(
                                list_value_range[0]) + ")"
                            output.cell(row=out_row, column=3).value = in_data.cell(row=data_row, column=3).value
                            output.cell(row=out_row, column=4).value = in_data.cell(row=data_row, column=4).value
                            output.cell(row=out_row, column=5).value = in_data.cell(row=data_row, column=5).value
                            output.cell(row=out_row, column=6).value = in_data.cell(row=data_row, column=6).value
                            output.cell(row=out_row, column=7).value = in_data.cell(row=data_row, column=7).value
                            output.cell(row=out_row, column=8).value = in_data.cell(row=data_row, column=8).value
                            output.cell(row=out_row, column=9).value = in_data.cell(row=data_row, column=9).value
                            output.cell(row=out_row, column=10).value = "1. SW run normal \r\n 2. Input = " + str(
                                list_value_range[0])
                            output.cell(row=out_row, column=11).value = in_data.cell(row=data_row, column=11).value
                            output.cell(row=out_row, column=12).value = in_data.cell(row=data_row,
                                                                                     column=12).value + " (" + str(
                                list_value_range[0]) + ")"
                            out_row = out_row + 1

                    elif count_value_range == 2:
                        for i in (0, 1):
                            output.cell(row=out_row, column=1).value = in_data.cell(row=data_row, column=1).value
                            output.cell(row=out_row, column=2).value = in_data.cell(row=data_row,
                                                                                    column=2).value + " (" + str(
                                list_value_range[i]) + ")"
                            output.cell(row=out_row, column=3).value = in_data.cell(row=data_row, column=3).value
                            output.cell(row=out_row, column=4).value = in_data.cell(row=data_row, column=4).value
                            output.cell(row=out_row, column=5).value = in_data.cell(row=data_row, column=5).value
                            output.cell(row=out_row, column=6).value = in_data.cell(row=data_row, column=6).value
                            output.cell(row=out_row, column=7).value = in_data.cell(row=data_row, column=7).value
                            output.cell(row=out_row, column=8).value = in_data.cell(row=data_row, column=8).value
                            output.cell(row=out_row, column=9).value = in_data.cell(row=data_row, column=9).value
                            output.cell(row=out_row, column=10).value = "1. SW run normal \r\n 2. Input = " + str(
                                list_value_range[i])
                            output.cell(row=out_row, column=11).value = in_data.cell(row=data_row, column=11).value
                            output.cell(row=out_row, column=12).value = in_data.cell(row=data_row,
                                                                                     column=12).value + " (" + str(
                                list_value_range[i]) + ")"
                            out_row = out_row + 1
                    elif count_value_range >= 3:
                        for i in (0, 1, count_value_range - 1):
                            output.cell(row=out_row, column=1).value = in_data.cell(row=data_row, column=1).value
                            output.cell(row=out_row, column=2).value = in_data.cell(row=data_row,
                                                                                    column=2).value + " (" + str(
                                list_value_range[i]) + ")"
                            output.cell(row=out_row, column=3).value = in_data.cell(row=data_row, column=3).value
                            output.cell(row=out_row, column=4).value = in_data.cell(row=data_row, column=4).value
                            output.cell(row=out_row, column=5).value = in_data.cell(row=data_row, column=5).value
                            output.cell(row=out_row, column=6).value = in_data.cell(row=data_row, column=6).value
                            output.cell(row=out_row, column=7).value = in_data.cell(row=data_row, column=7).value
                            output.cell(row=out_row, column=8).value = in_data.cell(row=data_row, column=8).value
                            output.cell(row=out_row, column=9).value = in_data.cell(row=data_row, column=9).value
                            output.cell(row=out_row, column=10).value = "1. SW run normal \r\n 2. Input = " + str(
                                list_value_range[i])
                            output.cell(row=out_row, column=11).value = in_data.cell(row=data_row, column=11).value
                            output.cell(row=out_row, column=12).value = in_data.cell(row=data_row,
                                                                                     column=12).value + " (" + str(
                                list_value_range[i]) + ")"
                            out_row = out_row + 1

                elif (type_defined == True) and (range_defined == False):
                    # Write Type is define to ListTypeNotDef but not range
                    if (range_defined == False):
                        list2.cell(row=out_row3, column=1).value = in_data.cell(row=data_row, column=13).value
                        out_row3 = out_row3 + 1

        else:
            # Write test case with normal type (CS, Mess,)
            output.cell(row=out_row, column=1).value = in_data.cell(row=data_row, column=1).value
            output.cell(row=out_row, column=2).value = in_data.cell(row=data_row, column=2).value
            output.cell(row=out_row, column=3).value = in_data.cell(row=data_row, column=3).value
            output.cell(row=out_row, column=4).value = in_data.cell(row=data_row, column=4).value
            output.cell(row=out_row, column=5).value = in_data.cell(row=data_row, column=5).value
            output.cell(row=out_row, column=6).value = in_data.cell(row=data_row, column=6).value
            output.cell(row=out_row, column=7).value = in_data.cell(row=data_row, column=7).value
            output.cell(row=out_row, column=8).value = in_data.cell(row=data_row, column=8).value
            output.cell(row=out_row, column=9).value = in_data.cell(row=data_row, column=9).value
            output.cell(row=out_row, column=10).value = in_data.cell(row=data_row, column=10).value
            output.cell(row=out_row, column=11).value = in_data.cell(row=data_row, column=11).value
            output.cell(row=out_row, column=12).value = in_data.cell(row=data_row, column=12).value
            out_row = out_row + 1

    output_path = name_file + "_Test_Specification.xlsx"
    out.save(output_path)
    out.close()
    data.close()

def remove_files(path):
    os.remove(path)
    print("Removed file: ",path)

#Ham main xu ly moi thu
def SABTestSpecGen(file_path, name_file_path):
    _int_file_path = name_file_path + ".xlsx"
    _int_data_path = "_int_" + name_file_path + ".xlsx"
    _pre_output_path = "_preOut_" + name_file_path + ".xlsx"
    _list_rp = "_List_Report_" + name_file_path + ".xlsx"
    ##Xu ly du lieu dau vao
    if file_path.endswith('.xls'):
        print("End file is xls")
        cvt_xls_to_xlsx(file_path, name_file_path)
        initDataInput(_int_file_path, name_file_path)

    elif file_path.endswith('.xlsx'):
        print("End file is not xls, this is xlsx")
        initDataInput(file_path, name_file_path)

    #Read data
    readDataInput(_int_data_path, name_file_path)
    FinalOutput(_pre_output_path, name_file_path)

    #Tao report

    #Xoa file tam
    remove_files(_int_file_path)
    remove_files(_int_data_path)
    remove_files(_pre_output_path)
    #Khi nao viet function tao report, thi xe update
    remove_files(_list_rp)